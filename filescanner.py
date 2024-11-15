import os
import openpyxl
from datetime import datetime

def update_tracker(root_folder, tracker_file):
    # Load the existing workbook and select the active sheet
    wb = openpyxl.load_workbook(tracker_file)
    ws = wb.active

    # Read existing data into a dictionary for quick lookup
    existing_files = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[4] and row[3]:  # Assuming the file name is in the third column, file size is in the 4th and created date in the 5th
            key = (row[3], row[4])  # created_date + file size
            existing_files[key] = row
        elif row[4]:    # if only created date is there in the file, file size is empty
            key = (row[2], row[3])  # file name + created date is the key in this case
            existing_files[key] = row

    # Traverse the root folder and update the tracker
    current_time = datetime.now().strftime('%Y-%m-%d %H:%M')
    found_files = set()

    for foldername, subfolders, filenames in os.walk(root_folder):
        for filename in filenames:
            if filename == 'desktop.ini' or filename.startswith('.'):
                continue  # Skip hidden files and 'desktop.ini'

            file_path = os.path.join(foldername, filename)
            file_size = os.path.getsize(file_path)
            created_date = datetime.fromtimestamp(os.path.getctime(file_path)).strftime('%Y-%m-%d %H:%M:%S')
            modified_date = datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%Y-%m-%d %H:%M:%S')
            key = (created_date, file_size)
            found_files.add(key)

            status = "Active"
            if key in existing_files:
                # Update the existing row
                for row in ws.iter_rows(min_row=2):
                    if (row[3].value, row[4].value) == key:
                        if row[2].value != filename:
                            if row[1].value:
                                row[1].value += f", {row[2].value}"
                            else:
                                row[1].value = row[2].value
                        row[0].value = foldername
                        row[2].value = filename
                        row[5].value = modified_date
                        row[6].value = current_time
                        row[7].value = status
                        if foldername.endswith("könyvelve"):
                            row[9].value = "könyvelve"
                        elif row[9].value == "könyvelve":
                            row[9].value = ""
                        break
            else:
                # Check if a file with the same filename and created_date exists and update the size if missing
                filename_key = (filename, created_date)
                if filename_key in existing_files:  #and not existing_files[filename_key][4]:  # size column is empty
                    for row in ws.iter_rows(min_row=2):
                        if (row[2].value, row[3].value) == filename_key:
                            row[4].value = file_size
                            break
                elif not filename.startswith("tracker"):
                    # Append a new row
                    new_row = [foldername, "", filename, created_date, file_size, modified_date, current_time, status]
                    if foldername.endswith("könyvelve"):
                        new_row.append("könyvelve")
                    else:
                        new_row.append("")
                    ws.append(new_row)

    # Mark files as "Deleted" if they were not found in the current run
    for row in ws.iter_rows(min_row=2):
        if not row[2].value.startswith("tracker"):      # we exclude the tracker file, as its size will change always
            key = (row[3].value, row[4].value)
            if key not in found_files:
                row[7].value = "Deleted"
            elif not row[0].value.endswith("könyvelve") and row[9].value == "könyvelve":
                row[9].value = ""

    # Save the workbook
    wb.save(tracker_file)
    print(f'Excel file "{tracker_file}" updated successfully.')

root_folder = 'G:\\My Drive\\Konyveles\\'
tracker_file = os.path.join(root_folder, 'tracker.xlsx')
update_tracker(root_folder, tracker_file)
